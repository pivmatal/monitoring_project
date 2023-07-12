#!/home/ufk/ufk_venv/bin/python
import django
import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "monitoring_project.settings")
django.setup()

from datetime import datetime, timezone
import pytz
localtime = pytz.timezone("Asia/Krasnoyarsk")

from monitoring.models import UK, PIF
from django.db.models import Q

import time
import re
import shutil
import openpyxl

"""
ogrn = '1037702037680'
file_name = 'raiffeisen.html'
extractor = 'raiffeisen'

pif = UK.objects.filter(uk_ogrn = ogrn).first()
pifset = PIF.objects.filter(Q(pif_uk = pif) & ~Q(pif_checkpage=''))
"""
#file_ = open(f'/home/ufk/monitoring_project/templates/{file_name}', 'w')

import importlib
#module = importlib.import_module(f'extractors.{extractor}')

#for pif in pifset:
#    start_time = time.time()
#    ext = module.Extractor(pif.pif_checkpage)
#    ext.scrape()
#    content = ext.get_data(720)
#    print(f'Execution time: {time.time() - start_time}')
#    file_.write('<div class="container">')
#    file_.write('<h3>{}</h3><a href="{}">{}</a>'.format(pif.pif_shortname, pif.pif_checkpage, pif.pif_checkpage))
#    for e in content:
#        if len(e) > 0:
#            file_.write('<h5>{}</h5>'.format(e[0]))
#            for k in e[1]:
#                if len(k) > 0:
#                    file_.write('<div class="d-flex flex-wrap justify-content-between">')
#                    file_.write('<a href=\"{}\">{}</a><p><b>{}</b>'.format(k[2], k[0], k[1]))
#                    file_.write('</div>')
#    file_.write('</div>')
#file_.close()

def filter_out(content, rules, count=1):
    filtered_content = []
    rules_list = rules.split(';;;')
    for section in content:
        if any([re.match(rule, section[0], re.IGNORECASE) for rule in rules_list]):
            filtered_content += section[1]
            continue
        filtered_content += [doc for doc in section[1] if any([re.match(rule, doc[0], re.IGNORECASE) for rule in rules_list])]
    return sorted(filtered_content, key=lambda tup: tup[1], reverse=True)[:count]

def write_line(ws, row, data):
    try:
        npp = ws.cell(row=row-1, column=1).value + 1
    except:
        npp = 1
    ws.cell(row=row, column=1).value = npp
    for i in range(6):
        ws.cell(row=row, column=2+i).value = data[i]

#uks = UK.objects.filter(~Q(uk_extractor = ''))
uks = UK.objects.all()
sum = 0
start_time = time.time()
report_data = []
total_start = time.time()
localtime = pytz.timezone("Asia/Krasnoyarsk")
date_now = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=localtime)
date_str = datetime.strftime(date_now, '%Y%m%d')
filename = f'/home/ufk/monitoring_project/static/report_kid_{date_str}.xlsx'
shutil.copy('/home/ufk/monitoring_project/static/templates/report_kid.xlsx', filename)
wb = openpyxl.load_workbook(filename)
ws = wb.active
wb.save(filename=filename)
for uk in uks:
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    current_row = ws.max_row + 1
    print(f'Doing UK: {uk.uk_name}')
    start_time = time.time()
    module = None
    if uk.uk_extractor:
        module = importlib.import_module(f'extractors.{uk.uk_extractor}')
#    for pif in PIF.objects.filter(Q(pif_uk = uk) & ~Q(pif_checkpage='')):
    for pif in PIF.objects.filter(Q(pif_uk = uk)):
        ext = None
        print(pif.pif_name)
        if not pif.pif_enabled:
            continue
        if pif.pif_checkpage and module:
            result = []
            try:
                ext = module.Extractor(pif.pif_checkpage)
                ext.scrape()
                content = filter_out(ext.get_data(13), '.*КИД.*;;;.*ключ.*инф.*доку', 5)
                for doc in content:
                    report_data.append((pif.pif_npp, pif.pif_name, '://'.join((uk.uk_sitetype, uk.uk_site)), \
                            doc[0], doc[1].strftime("%d.%m.%Y %H:%M"), doc[2]))
                    result = [pif.pif_uk.uk_name, pif.pif_name, pif.pif_checkpage, doc[0], doc[2], doc[1].strftime("%d.%m.%Y %H:%M")] 
                    write_line(ws, current_row, result)
                    current_row = current_row + 1
            except:
                report_data.append((pif.pif_npp, pif.pif_name, '://'.join((uk.uk_sitetype, uk.uk_site)), \
                    "", "", ""))
                print(f'Ошибка при получении: {pif.pif_name} \t {pif.pif_checkpage}')
            if len(result) == 0:
                result = [pif.pif_uk.uk_name, pif.pif_name, pif.pif_checkpage, '', '', '']
                write_line(ws, current_row, result)
                current_row = current_row + 1
        else:
            report_data.append((pif.pif_npp, pif.pif_name, '://'.join((uk.uk_sitetype, uk.uk_site)), \
                    "", "", ""))

        if pif.pif_checkpage.strip() == '':
            result = [pif.pif_uk.uk_name, pif.pif_name, pif.pif_checkpage, '', '', '']
            write_line(ws, current_row, result)
            current_row = current_row + 1
        
        if len(result) == 0:
            result = [pif.pif_uk.uk_name, pif.pif_name, pif.pif_checkpage, '', '', '']
            write_line(ws, current_row, result)
            current_row = current_row + 1

        if module and ext:
            del ext
    wb.save(filename)
    del module
    print(f'UK execution time: {time.time() - start_time}')
#for row in report_data:
#    ws.append(row)
#wb.save(filename=filename)    
print(f'Total execution time: {time.time() - total_start}')
