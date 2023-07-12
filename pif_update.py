#!/home/ufk/ufk_venv/bin/python
import requests
import pandas as pd
import numpy as np

import django
from django.db.models import Q

import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "monitoring_project.settings")

django.setup()

from monitoring.models import UK, PIF, CheckLog

import pytz
from datetime import datetime, timezone, timedelta

import shutil
import importlib
import re
import openpyxl

localtime = pytz.timezone("Asia/Krasnoyarsk")
date_now = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=localtime)
date_str = datetime.strftime(date_now, '%Y%m%d')
report_file = f'/home/ufk/monitoring_project/static/MonitoringPIF_{date_str}.xlsx'

def differentiator(row):
    keys_loc = {key: df.columns.get_loc(key) for key in keys + new_keys}
    diff_list = []
    if row[keys_loc['pif_npp']] == row[keys_loc['pif_npp_r']]:
        for key in keys[1:]:
            if row[keys_loc[key]] != row[keys_loc[f'{key}_r']]:
                print(key)
                print()
                print(row[keys_loc[key]])
                print()
                print(row[keys_loc[f'{key}_r']])
                diff_list.append(key)
    if len(diff_list) > 0:
        return ';'.join(diff_list)
    else:
        return np.nan


def insert_pifs(df):
    pif_statuses = {
        1 : 'Зарегистрирован',
        2 : 'Формируется',
        3 : 'Сформирован',
        4 : 'В стадии прекращения',
        5 : 'Истёк срок формирования',
        6 : 'Исключён из реестра',
        7 : 'Согласован',
        } 
    df['pif_status'] = df['pif_status'].map(pif_statuses)
    output_df = pd.DataFrame()
    for index, row in df.iterrows():
        try:
            uk = UK.objects.filter(uk_ogrn = row[17]).first()
        except UK.DoesNotExist:
            uk = None
        if not uk:
            try:
                uk = UK.objects.filter(uk_name = row[16].strip()).first()
            except UK.DoesNotExist:
                uk = None
        if uk:
            ins = PIF()
            ins.pif_uk = uk
            ins.pif_enabled = True
            for key in db_fields.values():
                setattr(ins, key, row[key])
            try:
                ins.save()
                output_df = pd.concat([output_df, df.loc[index].to_frame().T])
            except Exception as e:
                print(str(e))
    return output_df

def update_pifs(df):
    pif_statuses = {
        1 : 'Зарегистрирован',
        2 : 'Формируется',
        3 : 'Сформирован',
        4 : 'В стадии прекращения',
        5 : 'Истёк срок формирования',
        6 : 'Исключён из реестра',
        7 : 'Согласован',
        } 
    df['pif_status'] = df['pif_status'].map(pif_statuses)
    for index, row in df.iterrows():
        requisites = []
        if row['diff'] is not np.nan:
            requisites = row['diff'].split(';')
        if len(requisites) > 0:
            try:
                pif = PIF.objects.get(pif_npp = row['pif_npp'])
                if pif:
                    for e in requisites:
                        print(getattr(pif, e) == row[e])
                        setattr(pif, e, row[e])
                    pif.save()
            except PIF.DoesNotExist:
                print(f'Не найден ПИФ: {row["pif_npp"]}')

def filter_out(content, rules, count=1):
    filtered_content = []
    rules_list = rules.split(';;;')
    for section in content:
        if any([re.match(rule, section[0], re.IGNORECASE) for rule in rules_list]):
            filtered_content += section[1]
            continue
        filtered_content += [doc for doc in section[1] if any([re.match(rule, doc[0], re.IGNORECASE) for rule in rules_list])]
    return sorted(filtered_content, key=lambda tup: tup[1], reverse=True)[:count]

def log_changes(df):
    for index, row in df.iterrows():
        try:
            pif = PIF.objects.get(pif_npp = row['pif_npp'])
            acknowledged = True
            text = [] 
            if (row['diff'] is not np.nan):
                for key in row['diff'].split(';'):
                    text.append(f'Графа {df.columns.get_loc(key)}: {getattr(row, key, "").split(",")[-1]}')
                if ('pif_npdu' in row['diff']) or ('pif_prefpdu' in row['diff']):
                    acknowledged = False
            elif row['diff'] is np.nan:
                for key in ['pif_npdu', 'pif_dpdu']:
                    text.append(f'Графа {df.columns.get_loc(key)+1}: {getattr(row, key, "")}')
            text = '\n'.join(text)
            doc_name = ''
            doc_date = None
            error = None
            if pif.pif_checkpage.strip() != '':
                try:
                    module = importlib.import_module(f'extractors.{pif.pif_uk.uk_extractor}')
                    ext = module.Extractor(pif.pif_checkpage)
                    ext.scrape()
                    content = filter_out(ext.get_data(1), '.*ПДУ.*;;;.*правил.*ДУ.*;;;.*правил.*довер.*упр.*;;;.*измен.*правил.*')
                    if len(content) > 0:
                        doc = content[0]
                        doc_name = f'{doc[0]} ({doc[2]})'
                        doc_date = doc[1]
                    del ext
                except:
                    error = 'Ошибка при проверке сайта. Проверьте работу правил для УК!'
            else:
                error = 'Страница раскрытия информации не задана!'
            ins = CheckLog(pif = pif, text = text, doc_name = doc_name, 
                            doc_date = doc_date, error = error, 
                            acknowledged = acknowledged)
            ins.save()
            #print(pif, text, doc_name, doc_date, error)
        except PIF.DoesNotExist:
            print(f'Не найден ПИФ: {row["pif_npp"]}')

def write_report():
    shutil.copy('/home/ufk/monitoring_project/MonitoringPIF.xlsx', report_file)
    wb = openpyxl.load_workbook(report_file)
    ws = wb.active
    ws['E1'] = f'Сформировано: {datetime.now(tz=localtime).strftime("%d.%m.%Y")}'
    ten_days = (datetime.now(tz=localtime)-timedelta(days=10)).replace(hour=0, minute=0, second=0)
    today = datetime.now(tz=localtime).replace(hour=0, minute=0, second=0) 
    today_logs = CheckLog.objects.filter(Q(date_log__gte = today) & Q(pif__isnull = False))
    current_row = 5
    npp = 1
    for element in today_logs:
    #for element in []:
        ws.cell(row=current_row, column=1).value = npp
        ws.cell(row=current_row, column=2).value = element.pif.pif_status
        ws.cell(row=current_row, column=3).value = re.sub('<.*$', '', element.pif.pif_checkpage)
        ws.cell(row=current_row, column=4).value = element.pif.pif_uk.uk_shortname
        ws.cell(row=current_row, column=5).value = element.pif.pif_shortname
        if element.pif.pif_prefpdu:
            ws.cell(row=current_row, column=6).value = element.pif.pif_prefpdu.split(",")[-1] 
        else:
            ws.cell(row=current_row, column=6).value = f'{element.pif.pif_npdu} зарегистрированы {element.pif.pif_dpdu}'
        ws.cell(row=current_row, column=7).value = element.text
        ws.cell(row=current_row, column=8).value = element.doc_name
        if element.doc_date:
            ws.cell(row=current_row, column=9).value = element.doc_date.strftime("%d.%m.%Y %H:%M")
            ws.cell(row=current_row, column=10).value = 'OK'
        if element.error:
            ws.cell(row=current_row, column=10).value = element.error
        current_row += 1
        npp += 1
    old_logs = CheckLog.objects.filter(Q(date_log__gte=ten_days) & Q(date_log__lte=today)
    #old_logs = CheckLog.objects.filter(Q(date_log__gte=ten_days)
            & Q(pif__isnull = False) & Q(doc_name = '') & Q(acknowledged = False))
    for element in old_logs:
        ws.cell(row=current_row, column=1).value = npp
        ws.cell(row=current_row, column=2).value = element.pif.pif_status
        ws.cell(row=current_row, column=3).value = re.sub('<.*$', '', element.pif.pif_checkpage)
        ws.cell(row=current_row, column=4).value = element.pif.pif_uk.uk_shortname
        ws.cell(row=current_row, column=5).value = element.pif.pif_shortname
        if element.pif.pif_prefpdu:
            ws.cell(row=current_row, column=6).value = element.pif.pif_prefpdu.split(",")[-1] 
        else:
            ws.cell(row=current_row, column=6).value = f'{element.pif.pif_npdu} зарегистрированы {element.pif.pif_dpdu}'
        if element.pif.pif_checkpage.strip() != '':
            try:
                module = importlib.import_module(f'extractors.{element.pif.pif_uk.uk_extractor}')
                ext = module.Extractor(element.pif.pif_checkpage)
                ext.scrape()
                num_of_days = datetime.now().astimezone(localtime) - element.date_log.astimezone(localtime) 
                content = filter_out(ext.get_data(num_of_days.days + 5), '.*ПДУ.*;;;.*правил.*ДУ.*;;;.*правил.*довер.*упр.*;;;.*измен.*правил.*')
                print(content)
                if len(content) > 0:
                    doc = content[0]
                    setattr(element, 'doc_name', f'{doc[0]} ({doc[2]})')
                    setattr(element, 'doc_date', doc[1])
                    setattr(element, 'acknowledged', True)
                    setattr(element, 'error', '')
                    element.save()
                del ext
            except Exception as e:
                print(e)
                pass
        ws.cell(row=current_row, column=7).value = element.text
        ws.cell(row=current_row, column=8).value = element.doc_name
        if element.doc_date:
            ws.cell(row=current_row, column=9).value = element.doc_date.strftime("%d.%m.%Y %H:%M")
            ws.cell(row=current_row, column=10).value = 'OK'
        if element.error:
            ws.cell(row=current_row, column=10).value = element.error
        current_row += 1
        npp += 1
    wb.save(report_file)    

pif_url = "https://cbr.ru/vfs/finmarkets/files/supervision/list_PIF.xlsx"
req = requests.get(pif_url, allow_redirects=True)
pif_file = open("pif.xlsx", "wb").write(req.content)
########################
pifs = pd.read_excel('pif.xlsx', skiprows=2)
for col in pifs.select_dtypes('object').columns:
    pifs[col] = pifs[col].str.strip()
qual_name = 'Инвестиционные паи предназначены для квалифицированного инвестора'
status_name = 'Статус'
db_fields = {
        0: 'pif_npp',
        1: 'pif_type',
        2: 'pif_status',
        3: 'pif_name',
        4: 'pif_shortname',
        5: 'pif_category',
        6: 'pif_npdu',
        7: 'pif_dpdu',
        8: 'pif_spdu',
        14: 'pif_prefnames',
        15: 'pif_prefpdu',
        }
pif_statuses = {
        'Зарегистрирован': 1,
        'Формируется': 2,
        'Сформирован': 3,
        'В стадии прекращения': 4,
        'Истёк срок формирования': 5,
        'Исключён из реестра': 6,
        'Согласован': 7,
        }
pifs[status_name] = pifs[status_name].map(pif_statuses)
pifs[qual_name] = pifs[qual_name].map({'да': 1, '': 0})
django_df = pd.DataFrame(list(PIF.objects.all().values()))
django_df['pif_status'] = django_df['pif_status'].map(pif_statuses)
columns = {pifs.columns[field]: db_fields[field] for field in db_fields.keys()}
pifs.rename(columns, axis=1, inplace=True)
df = pd.concat([pifs, django_df], ignore_index=True)
df.reset_index(inplace=True)
df = df.rename_axis('MyIdx').sort_values(by = ['pif_npp', 'MyIdx'], ascending = [False, True])
keys = [db_fields[key] for key in db_fields.keys()]
new_keys = [f'{key}_r' for key in keys]
df[new_keys] = df[keys].shift(-1)

df['diff'] = df.apply(differentiator, axis=1)
#print(df.loc[df['diff'].notna()])
#---------------------
filtered_pifs = pifs[(pifs[qual_name] == 0) & (pifs['pif_status'] < 6)]
new_df = insert_pifs(filtered_pifs[~filtered_pifs['pif_npp'].isin(django_df['pif_npp'].to_list())])
df = pd.concat([df.loc[df['diff'].notna()], new_df])
print(df)
update_pifs(df)
log_changes(df)
write_report()
