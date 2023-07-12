#!/usr/bin/env python
from openpyxl import load_workbook

import django

import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "monitoring_project.settings")

django.setup()


from monitoring.models import PUCB

pucb_wb = load_workbook("pucb.xlsx")
pucb_ws = pucb_wb.worksheets[0]

added = 0
changed = 0
print("pucb update")
for row in pucb_ws.iter_rows(min_row = 2, values_only=True):
    ins = PUCB(pucb_npp = row[0],
                pucb_name = row[1],
                pucb_inn = row[2],
                pucb_sitetype = row[4],
                pucb_site = row[5])
    try:
        ins.save()
        added += 1
        print("Added {}".format(ins.pucb_site))
    except Exception as e:
        print(str(e))

print("pucb update ended:")
print("Added {} objects".format(added))
