#!/usr/bin/env python
import requests
from openpyxl import load_workbook

import django

import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "monitoring_project.settings")

django.setup()


from monitoring.models import UK
from monitoring.models import PIF

uk_url = "https://cbr.ru/vfs/finmarkets/files/supervision/list_managementcompanies.xlsx"

req = requests.get(uk_url, allow_redirects=True)
uk_file = open("uk.xlsx", "wb").write(req.content)

uk_wb = load_workbook("uk.xlsx")
uk_ws = uk_wb.worksheets[0]

added = 0
changed = 0
print("UKs site update started...")
for row in uk_ws.iter_rows(min_row = 2, values_only=True):
    
    try:
        uk = UK.objects.get(uk_ogrn = row[4])

    except UK.DoesNotExist as dne:
            
        if row[10].strip() != '':
            site = row[10]
            if site.find("http://") != -1:
                sitetype = "http"
            else:
                sitetype = "https"
        else:
            site = ""
            sitetype = ""
        
        site = site.replace("http://", "").replace("https://", "")
        ins = UK(uk_npp = row[0],
                uk_name = row[1],
                uk_shortname = row[2],
                uk_inn = row[3],
                uk_ogrn = row[4],
                uk_nlic = row[5],
                uk_dlic = row[6],
                uk_slic = row[7],
                uk_addr = row[8],
                uk_phones = row[9],
                uk_site = site,
                uk_sitetype = sitetype,
                uk_enabled = True)
        try:
            ins.save()
            added += 1
            print("Added {}".format(ins))
        except Exception as e:
            print(str(e))

    else:
        need_change = 0
        fields = {
                'uk_npp': 0,
                'uk_name': 1,
                'uk_shortname': 2,
                'uk_inn': 3,
                'uk_ogrn': 4,
                'uk_nlic': 5,
                'uk_dlic': 6,
                'uk_slic': 7,
                'uk_addr': 8,
                'uk_phones': 9,
                }
        for key in fields.keys():
            if getattr(uk, key) != row[fields[key]]:
                setattr(uk, key, row[fields[key]])
                need_change = 1
        
        if uk.uk_site != row[10].strip().replace("http://", "").replace("https://", ""):
            site = row[10] 
            if site.find("http://") != -1:
                sitetype = "http"
            else:
                sitetype = "https"
            site = site.replace("http://", "").replace("https://", "")
            uk.uk_site = site
            uk.uk_sitetype = sitetype
            need_change = 1

        if need_change != 0:
            uk.save()
print("UKs site update ended:")
print("Added {} objects".format(added))
print("Changed {} objects".format(changed))

pif_url = "https://cbr.ru/vfs/finmarkets/files/supervision/list_PIF.xlsx"
req = requests.get(pif_url, allow_redirects=True)
pif_file = open("pif.xlsx", "wb").write(req.content)

pif_wb = load_workbook("pif.xlsx")
pif_ws = pif_wb.worksheets[0]

added = 0
changed = 0
fields = {
        'pif_npp': 0,
        'pif_type': 1,
        'pif_status': 2,
        'pif_name': 3,
        'pif_shortname': 4,
        'pif_category': 5,
        'pif_npdu': 6,
        'pif_dpdu': 7,
        'pif_spdu': 8,
        'pif_prefnames': 14,
        'pif_prefpdu': 15,
        }

print("\nPIFs update started")

for row in pif_ws.iter_rows(min_row = 4, values_only=True):
    
    try:
        pif = PIF.objects.get(pif_name = row[3])

    except PIF.DoesNotExist:
            
        if row[2].strip().lower() != "Исключён из реестра".lower() and str(row[10]).strip() == '':
            if str(row[17]).strip() != '':
                try:
                    uk = UK.objects.get(uk_ogrn = row[17])
                except UK.DoesNotExist:
                    print("N{} UK with orgn {} did not found!".format(row[0], row[17]))
                    uk = None
            else:
                try:
                    uk = UK.objects.get(uk_name = row[16].strip())
                except UK.DoesNotExist:
                    print("N{} UK with name {} did not found!".format(row[0], row[16]))
                    uk = None
            if uk is None:
                print("N{} PIF has no UK found!".format(row[0], row[16]))
            else:
                ins = PIF()
                for key in fields.keys():
                    setattr(ins, key, row[fields[key]])
                ins.pif_uk = uk
                ins.pif_enabled = True
                try:
                    ins.save()
                    added += 1
                    print("Added {}".format(ins))
                except Exception as e:
                    print(str(e))

    else:
        need_change = 0
        for key in fields.keys():
            if getattr(pif, key) != row[fields[key]]:
                setattr(pif, key, row[fields[key]])
                need_change = 1
        
        if need_change != 0:
            pif.save()
print("Added {} objects".format(added))
print("Changed {} objects".format(changed))


#req = requests.get(pif_url, allow_redirects=True)
#pif_file = open("pif.xlsx", "wb").write(req.content)
